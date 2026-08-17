"""商品管理面板"""
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton,
    QLabel, QLineEdit, QTextEdit, QTableWidget, QTableWidgetItem,
    QGroupBox, QFormLayout, QSpinBox, QHeaderView,
    QCheckBox, QDialog, QDialogButtonBox, QFileDialog
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QFont
from ui.msgbox import QMessageBox
import time
from models.product import Product


class ProductDialog(QDialog):
    """商品编辑对话框"""

    def __init__(self, parent=None, product: Product = None):
        super().__init__(parent)
        self.product = product or Product()
        self.setWindowTitle("编辑商品" if product and product.id else "添加商品")
        self.setMinimumWidth(450)
        self._build_ui()

    def _build_ui(self):
        layout = QFormLayout(self)
        layout.setSpacing(10)

        self.edit_name = QLineEdit(self.product.name)
        self.edit_name.setPlaceholderText("必填，如：儿童磁力积木")
        layout.addRow("商品名称：", self.edit_name)

        self.edit_price = QLineEdit(self.product.price)
        self.edit_price.setPlaceholderText("如：49元")
        layout.addRow("价格：", self.edit_price)

        self.edit_feature = QTextEdit()
        self.edit_feature.setFixedHeight(60)
        self.edit_feature.setPlainText(self.product.feature)
        self.edit_feature.setPlaceholderText("商品特点/卖点")
        layout.addRow("特点卖点：", self.edit_feature)

        self.edit_audience = QLineEdit(self.product.target_audience)
        self.edit_audience.setPlaceholderText("如：3-8岁儿童")
        layout.addRow("适合人群：", self.edit_audience)

        self.edit_benefit = QLineEdit(self.product.benefit)
        self.edit_benefit.setPlaceholderText("如：锻炼专注力，远离手机")
        layout.addRow("好处价值：", self.edit_benefit)

        self.edit_commission = QLineEdit(self.product.commission)
        self.edit_commission.setPlaceholderText("如：20%")
        layout.addRow("佣金：", self.edit_commission)

        self.edit_notes = QTextEdit()
        self.edit_notes.setFixedHeight(50)
        self.edit_notes.setPlainText(self.product.extra_notes)
        self.edit_notes.setPlaceholderText("补充话术信息，如：今天买一送一")
        layout.addRow("补充信息：", self.edit_notes)

        self.spin_scripts = QSpinBox()
        self.spin_scripts.setRange(1, 20)
        self.spin_scripts.setValue(self.product.scripts_per_round or 5)
        layout.addRow("每轮播报条数：", self.spin_scripts)

        # 按钮（汉化）
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.button(QDialogButtonBox.StandardButton.Ok).setText("确定")
        buttons.button(QDialogButtonBox.StandardButton.Cancel).setText("取消")
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addRow(buttons)

    def get_product(self) -> Product:
        p = self.product
        p.name = self.edit_name.text().strip()
        p.price = self.edit_price.text().strip()
        p.feature = self.edit_feature.toPlainText().strip()
        p.target_audience = self.edit_audience.text().strip()
        p.benefit = self.edit_benefit.text().strip()
        p.commission = self.edit_commission.text().strip()
        p.extra_notes = self.edit_notes.toPlainText().strip()
        p.scripts_per_round = self.spin_scripts.value()
        if not p.created_at:
            p.created_at = time.time()
        return p


class ProductPanel(QWidget):
    """商品管理页面"""

    def __init__(self, main_window):
        super().__init__()
        self.main_window = main_window
        self._build_ui()
        self.refresh_list()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)

        # 标题和按钮
        top_layout = QHBoxLayout()
        title = QLabel("商品管理")
        title.setFont(QFont("Microsoft YaHei", 14, QFont.Weight.Bold))
        top_layout.addWidget(title)
        top_layout.addStretch()

        self.btn_add = QPushButton("+ 添加商品")
        self.btn_add.setFixedHeight(38)
        self.btn_add.setProperty("class", "primary")
        self.btn_add.clicked.connect(self._on_add)
        top_layout.addWidget(self.btn_add)

        self.btn_edit = QPushButton("编辑")
        self.btn_edit.setFixedHeight(38)
        self.btn_edit.clicked.connect(self._on_edit)
        top_layout.addWidget(self.btn_edit)

        self.btn_delete = QPushButton("删除")
        self.btn_delete.setFixedHeight(38)
        self.btn_delete.setProperty("class", "text-danger")
        self.btn_delete.clicked.connect(self._on_delete)
        top_layout.addWidget(self.btn_delete)

        self.btn_import = QPushButton("Excel导入")
        self.btn_import.setFixedHeight(38)
        self.btn_import.clicked.connect(self._on_import_excel)
        top_layout.addWidget(self.btn_import)

        self.btn_fetch_cart = QPushButton("拉取小黄车")
        self.btn_fetch_cart.setFixedHeight(38)
        self.btn_fetch_cart.setProperty("class", "warning")
        self.btn_fetch_cart.clicked.connect(self._on_fetch_cart)
        top_layout.addWidget(self.btn_fetch_cart)

        layout.addLayout(top_layout)

        # 商品表格
        self.table = QTableWidget()
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels(["名称", "价格", "特点", "人群", "话术数", "启用"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.doubleClicked.connect(self._on_edit)
        layout.addWidget(self.table)

    def refresh_list(self):
        """刷新商品列表"""
        products = self.main_window.db.get_all_products()
        self.table.setRowCount(len(products))
        for row, p in enumerate(products):
            scripts = self.main_window.db.get_scripts_by_product(p.id, "main")
            self.table.setItem(row, 0, QTableWidgetItem(p.name))
            self.table.setItem(row, 1, QTableWidgetItem(p.price))
            self.table.setItem(row, 2, QTableWidgetItem(p.feature[:30]))
            self.table.setItem(row, 3, QTableWidgetItem(p.target_audience))
            self.table.setItem(row, 4, QTableWidgetItem(str(len(scripts))))
            chk = QTableWidgetItem("✓" if p.enabled else "✗")
            chk.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(row, 5, chk)
            # 存储product id
            self.table.item(row, 0).setData(Qt.ItemDataRole.UserRole, p.id)

    def _get_selected_product(self) -> Product:
        row = self.table.currentRow()
        if row < 0:
            return None
        pid = self.table.item(row, 0).data(Qt.ItemDataRole.UserRole)
        products = self.main_window.db.get_all_products()
        for p in products:
            if p.id == pid:
                return p
        return None

    def _on_add(self):
        dlg = ProductDialog(self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            product = dlg.get_product()
            if not product.name:
                QMessageBox.warning(self, "提示", "商品名称不能为空")
                return
            self.main_window.db.add_product(product)
            self.refresh_list()

    def _on_edit(self):
        product = self._get_selected_product()
        if not product:
            QMessageBox.information(self, "提示", "请先选择一个商品")
            return
        dlg = ProductDialog(self, product)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            updated = dlg.get_product()
            self.main_window.db.update_product(updated)
            self.refresh_list()

    def _on_delete(self):
        product = self._get_selected_product()
        if not product:
            return
        ret = QMessageBox.question(self, "确认删除", f"确定删除商品「{product.name}」及其所有话术？")
        if ret == QMessageBox.StandardButton.Yes:
            self.main_window.db.delete_product(product.id)
            self.refresh_list()

    def _on_import_excel(self):
        """Excel批量导入商品"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择Excel文件", "",
            "Excel文件 (*.xlsx *.xls);;CSV文件 (*.csv)"
        )
        if not file_path:
            return

        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb.active

            # 读取表头确定列顺序
            headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]

            # 列名映射（支持多种常见表头）
            col_map = {}
            name_keys = ["名称", "商品名", "商品名称", "name", "产品"]
            price_keys = ["价格", "售价", "price", "单价"]
            feature_keys = ["特点", "卖点", "feature", "描述", "简介"]
            audience_keys = ["人群", "适合", "年龄", "audience", "目标"]
            benefit_keys = ["好处", "价值", "benefit", "功效"]

            def find_col(keys):
                for i, h in enumerate(headers):
                    if any(k in h.lower() for k in keys):
                        return i
                return -1

            name_col = find_col(name_keys)
            price_col = find_col(price_keys)
            feature_col = find_col(feature_keys)
            audience_col = find_col(audience_keys)
            benefit_col = find_col(benefit_keys)

            if name_col < 0:
                # 如果没有表头，默认第一列是名称
                name_col = 0

            count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[name_col]:
                    continue
                name = str(row[name_col]).strip()
                if not name:
                    continue

                product = Product(
                    name=name,
                    price=str(row[price_col]).strip() if price_col >= 0 and price_col < len(row) and row[price_col] else "",
                    feature=str(row[feature_col]).strip() if feature_col >= 0 and feature_col < len(row) and row[feature_col] else "",
                    target_audience=str(row[audience_col]).strip() if audience_col >= 0 and audience_col < len(row) and row[audience_col] else "",
                    benefit=str(row[benefit_col]).strip() if benefit_col >= 0 and benefit_col < len(row) and row[benefit_col] else "",
                    created_at=time.time(),
                )
                self.main_window.db.add_product(product)
                count += 1

            wb.close()
            self.refresh_list()
            QMessageBox.information(self, "导入成功", f"成功导入 {count} 个商品！")

        except Exception as e:
            QMessageBox.critical(self, "导入失败", f"错误：{str(e)}\n\n请确保Excel第一行是表头，包含'名称'列。")

    def _on_fetch_cart(self):
        """拉取直播间小黄车商品"""
        from ui.msgbox import QInputDialog
        from core.douyin_connector import DouyinConnector

        # 优先使用弹幕面板已配置的直播间URL
        default_url = self.main_window.config.danmaku.room_url or ""
        url, ok = QInputDialog.getText(
            self, "拉取小黄车商品",
            "输入直播间URL或房间号：",
            text=default_url
        )
        if not ok or not url.strip():
            return

        # 检查是否已有cookie，没有则弹出登录窗口
        cookie = self.main_window.douyin_connector.cookie
        if not cookie:
            reply = QMessageBox.question(
                self, "需要登录",
                "拉取小黄车需要抹音登录态。\n\n是否现在登录抹音？",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.Yes
            )
            if reply == QMessageBox.StandardButton.Yes:
                self._open_douyin_login()
                cookie = self.main_window.douyin_connector.cookie
            if not cookie:
                return

        self.btn_fetch_cart.setEnabled(False)
        self.btn_fetch_cart.setText("拉取中...")
        self._pending_fetch_url = url.strip()

        try:
            products = DouyinConnector.fetch_room_products(self._pending_fetch_url, cookie=cookie)
            self._save_fetched_products(products)
            self.btn_fetch_cart.setEnabled(True)
            self.btn_fetch_cart.setText("拉取小黄车")
        except Exception as e:
            # HTTP方式失败（可能被反爬签名拦截），自动尝试浏览器方式
            self.btn_fetch_cart.setEnabled(True)
            self.btn_fetch_cart.setText("拉取小黄车")
            reply = QMessageBox.question(
                self, "HTTP拉取失败",
                f"{str(e)}\n\n是否使用内嵌浏览器方式拉取？\n（在页面中调用接口，自动携带登录态和签名，成功率更高）",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.Yes
            )
            if reply == QMessageBox.StandardButton.Yes:
                self._fetch_via_browser(self._pending_fetch_url)

    def _save_fetched_products(self, products):
        """保存拉取到的商品到数据库（自动填充销量/原价作为初始卖点）"""
        count = 0
        for item in products:
            # 用拉取到的富字段自动生成一段卖点，方便后续AI话术生成
            notes = []
            sales = str(item.get("sales", "") or "").strip()
            if sales:
                notes.append(f"已售{sales}件")
            market_price = str(item.get("market_price", "") or "").strip()
            if market_price and market_price != item.get("price", ""):
                notes.append(f"原价{market_price}")
            feature = "｜".join(notes)

            product = Product(
                name=item["name"],
                price=item.get("price", ""),
                feature=feature,
                created_at=time.time(),
            )
            self.main_window.db.add_product(product)
            count += 1
        self.refresh_list()
        QMessageBox.information(self, "拉取成功", f"成功从小黄车拉取 {count} 个商品！")

    def _fetch_via_browser(self, url: str):
        """使用内嵌浏览器拉取商品（处理反爬签名）"""
        try:
            from ui.product_fetch_dialog import ProductFetchDialog, WEBENGINE_AVAILABLE
            if not WEBENGINE_AVAILABLE:
                QMessageBox.warning(
                    self, "缺少组件",
                    "内嵌浏览器组件未安装，请运行: pip install PyQt6-WebEngine"
                )
                return
            self._fetch_dialog = ProductFetchDialog(url, self)
            self._fetch_dialog.products_fetched.connect(self._on_browser_fetch_success)
            self._fetch_dialog.fetch_failed.connect(self._on_browser_fetch_failed)
            self._fetch_dialog.show()
            self._fetch_dialog.start_fetch()
        except ImportError as e:
            QMessageBox.warning(self, "缺少组件", f"内嵌浏览器组件未安装：{e}")

    def _on_browser_fetch_success(self, products):
        """浏览器拉取成功"""
        if self._fetch_dialog:
            self._fetch_dialog.close()
            self._fetch_dialog = None
        self._save_fetched_products(products)

    def _on_browser_fetch_failed(self, message: str):
        """浏览器拉取失败"""
        if self._fetch_dialog:
            self._fetch_dialog.close()
            self._fetch_dialog = None
        QMessageBox.warning(self, "拉取失败", message)

    def _open_douyin_login(self):
        """打开抹音登录窗口"""
        try:
            from ui.douyin_login_dialog import DouyinLoginDialog
            dlg = DouyinLoginDialog(self)
            dlg.login_success.connect(self._on_login_success)
            dlg.exec()
        except ImportError as e:
            QMessageBox.warning(
                self, "缺少组件",
                f"内嵌浏览器组件未安装：{e}\n\n"
                f"请运行: pip install PyQt6-WebEngine"
            )

    def _on_login_success(self, cookie: str):
        """登录成功回调"""
        self.main_window.douyin_connector.cookie = cookie
